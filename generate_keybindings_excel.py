#!/usr/bin/env python3
"""
ThinktankWebCC キーバインディング比較表 Excel生成スクリプト
目的: Editor/Table/WebViewモードの共通キーアサイン設計と空きキー把握
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# カラー定義
# ============================================================
C = {
    'header_bg':  '1F4E79',
    'header_fg':  'FFFFFF',
    'section_bg': '2E75B6',
    'section_fg': 'FFFFFF',
    'common3':    '70AD47',   # 3モード共通 - 緑
    'common2':    'FFD966',   # 2モード共通 - 黄
    'editor':     'BDD7EE',   # Editor専用 - 薄青
    'table':      'FCE4D6',   # Table専用  - 薄橙
    'webview':    'E2EFDA',   # WebView専用 - 薄緑
    'global':     'D9D9D9',   # グローバル  - 灰
    'mask':       'FFB3B3',   # MASK (NoAction) - 薄赤
    'exmode':     'FFF2CC',   # ExMode専用  - 薄黄
    'empty_e':    'DAEEFF',   # Editor空き  - 極薄青
    'empty_t':    'FFF0E8',   # Table空き   - 極薄橙
    'empty_w':    'F0FFF0',   # WebView空き - 極薄緑
    'empty':      'F5F5F5',   # 未使用      - 薄灰
    'white':      'FFFFFF',
}

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def fnt(bold=False, color='000000', size=9):
    return Font(bold=bold, color=color, size=size, name='Meiryo UI')

def aln(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def bdr(color='AAAAAA'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def h_cell(cell, value, bg=None, fg='FFFFFF', size=9, h='center'):
    bg = bg or C['header_bg']
    cell.value = value
    cell.fill = fill(bg)
    cell.font = fnt(bold=True, color=fg, size=size)
    cell.alignment = aln(h=h)
    cell.border = bdr()

def d_cell(cell, value, bg='FFFFFF', bold=False, size=9, color='000000', h='left'):
    cell.value = value
    cell.fill = fill(bg)
    cell.font = fnt(bold=bold, color=color, size=size)
    cell.alignment = aln(h=h)
    cell.border = bdr()

# ============================================================
# キーバインディングデータ
# Format: (context, mods, key, action_id, description_ja)
# Context: Panel-Mode-Tool-ExMode  (* = any)
# ============================================================
KEYBINDINGS = [
    # ---- MASK (Editor) ----
    ('*-Editor-Main-*', 'Alt',         'R',     'NoAction',                  '[MASK] 検索RegEx'),
    ('*-Editor-Main-*', 'Alt',         'C',     'NoAction',                  '[MASK] 検索大小文字'),
    ('*-Editor-Main-*', 'Alt',         'P',     'NoAction',                  '[MASK] 置換(大文字保持)'),
    ('*-Editor-Main-*', 'Control',     'H',     'NoAction',                  '[MASK] 置換'),
    # ---- MASK (Keyword) ----
    ('*-*-Keyword-*',   'Control',     'F',     'NoAction',                  '[MASK] 検索'),
    ('*-*-Keyword-*',   'Alt',         'R',     'NoAction',                  '[MASK] 検索RegEx'),
    ('*-*-Keyword-*',   'Alt',         'C',     'NoAction',                  '[MASK] 検索大小文字'),
    ('*-*-Keyword-*',   'Control',     'H',     'NoAction',                  '[MASK] 置換'),
    ('*-*-Keyword-*',   'Alt',         'P',     'NoAction',                  '[MASK] 置換(大文字保持)'),

    # ---- Global: Application delegate ----
    ('*-*-*-*', '',              'F5',   'App.Delegate',              'リロード'),
    ('*-*-*-*', 'Control+Shift', 'R',   'App.Delegate',              'リロード'),
    ('*-*-*-*', '',              'F12',  'App.Delegate',              '開発パネル'),
    ('*-*-*-*', '',              'F11',  'App.Delegate',              'フルスクリーン'),
    ('*-*-*-*', 'Control+Shift', '+',   'App.Delegate',              'フォント拡大'),
    ('*-*-*-*', 'Control+Shift', '=',   'App.Delegate',              'フォント縮小'),
    ('*-*-*-*', 'Control',       ';',   'App.Delegate',              'フォント拡大'),
    ('*-*-*-*', 'Control',       '-',   'App.Delegate',              'フォント縮小'),
    # ---- Global: Panel/Mode切替 ----
    ('*-*-*-*', 'Alt',           '\\',  'Panel:next',                '次パネル'),
    ('*-*-*-*', 'Alt+Shift',     '_',   'Panel:prev',                '前パネル'),
    ('*-*-*-*', 'Alt',           'L',   'Panel:Library',             'Libraryパネル'),
    ('*-*-*-*', 'Alt',           'I',   'Panel:Index',               'Indexパネル'),
    ('*-*-*-*', 'Alt',           'S',   'Panel:Shelf',               'Shelfパネル'),
    ('*-*-*-*', 'Alt',           'D',   'Panel:Desk',                'Deskパネル'),
    ('*-*-*-*', 'Alt',           '/',   'Panel:System',              'Systemパネル'),
    ('*-*-*-*', 'Alt',           '[',   'Panel:Chat',                'Chatパネル'),
    ('*-*-*-*', 'Alt',           ']',   'Panel:Log',                 'Logパネル'),
    ('*-*-*-*', 'Alt',           'Q',   'Mode:Table',                'Tableモードへ'),
    ('*-*-*-*', 'Alt',           'W',   'Mode:WebView',              'WebViewモードへ'),
    ('*-*-*-*', 'Alt',           'E',   'Mode:Editor',               'Editorモードへ'),
    ('*-*-*-*', 'Alt',           'M',   'Mode:next',                 '次モード'),
    ('*-*-*-*', 'Alt+Shift',     'M',   'Mode:prev',                 '前モード'),
    # ---- Global: ExMode起動 ----
    ('*-*-*-*', 'Alt+Control',   'A',   'ExMode:ExApp',              'ExAppモード'),
    ('*-*-*-*', 'Alt+Control',   'L',   'ExMode:ExLibrary',          'ExLibrary'),
    ('*-*-*-*', 'Alt+Control',   'I',   'ExMode:ExIndex',            'ExIndex'),
    ('*-*-*-*', 'Alt+Control',   'S',   'ExMode:ExShelf',            'ExShelf'),
    ('*-*-*-*', 'Alt+Control',   'D',   'ExMode:ExDesk',             'ExDesk'),
    ('*-*-*-*', 'Alt+Control',   '/',   'ExMode:ExSystem',           'ExSystem'),
    ('*-*-*-*', 'Alt+Control',   '[',   'ExMode:ExChat',             'ExChat'),
    ('*-*-*-*', 'Alt+Control',   ']',   'ExMode:ExLog',              'ExLog'),
    # ---- Global: Search / Misc ----
    ('*-*-*-*', 'Control+Alt',   'F',   'WebView.OpenSearch',        'グローバル検索'),
    ('*-*-*-*', 'Control',       'G',   'Editor.Memo.Create',        'メモ作成'),
    ('*-*-*-*', 'Alt',           'T',   'Editor.Date.Action',        'DateTime操作'),

    # ---- Editor: Edit/Move/Select ----
    ('*-Editor-Main-*', 'Control',       'Z',   'App.Delegate',              'Undo'),
    ('*-Editor-Main-*', 'Control',       'Y',   'App.Delegate',              'Redo'),
    ('*-Editor-Main-*', 'Control',       'X',   'App.Delegate',              '切り取り'),
    ('*-Editor-Main-*', 'Control',       'C',   'App.Delegate',              'コピー'),
    ('*-Editor-Main-*', 'Control',       'V',   'App.Delegate',              '貼り付け'),
    ('*-Editor-Main-*', '',              'UP',  'Editor.CurPos:prevline',    '前の行'),
    ('*-Editor-Main-*', '',              'DOWN','Editor.CurPos:nextline',    '次の行'),
    ('*-Editor-Main-*', '',              'LEFT','Editor.CurPos:prevchar',    '前の文字'),
    ('*-Editor-Main-*', '',              'RIGHT','Editor.CurPos:nextchar',   '次の文字'),
    ('*-Editor-Main-*', 'Control',       'P',   'Editor.CurPos:prevline',    '前の行(C-P)'),
    ('*-Editor-Main-*', 'Control',       'N',   'Editor.CurPos:nextline',    '次の行(C-N)'),
    ('*-Editor-Main-*', 'Control',       'F',   'Editor.CurPos:nextchar',    '次の文字(C-F)'),
    ('*-Editor-Main-*', 'Control',       'B',   'Editor.CurPos:prevchar',    '前の文字(C-B)'),
    ('*-Editor-Main-*', 'Control',       'A',   'Editor.CurPos:linestart+',  '行頭'),
    ('*-Editor-Main-*', 'Control',       'E',   'Editor.CurPos:lineend+',    '行末'),
    ('*-Editor-Main-*', 'Control+Alt',   'P',   'Editor.CurPos:firstline',   '先頭行'),
    ('*-Editor-Main-*', 'Control+Alt',   'N',   'Editor.CurPos:lastline',    '最終行'),
    ('*-Editor-Main-*', 'Alt',           'P',   'Editor.CurPos:prevvisfolding','前の見出し'),
    ('*-Editor-Main-*', 'Alt',           'N',   'Editor.CurPos:nextvisfolding','次の見出し'),
    ('*-Editor-Main-*', 'Shift',         'UP',  'Editor.SelPos:prevline',    '上へ選択'),
    ('*-Editor-Main-*', 'Shift',         'DOWN','Editor.SelPos:nextline',    '下へ選択'),
    ('*-Editor-Main-*', 'Shift',         'LEFT','Editor.SelPos:prevchar',    '左へ選択'),
    ('*-Editor-Main-*', 'Shift',         'RIGHT','Editor.SelPos:nextchar',   '右へ選択'),
    ('*-Editor-Main-*', 'Control+Shift', 'P',   'Editor.SelPos:prevline',    '上へ選択(C-S-P)'),
    ('*-Editor-Main-*', 'Control+Shift', 'N',   'Editor.SelPos:nextline',    '下へ選択(C-S-N)'),
    ('*-Editor-Main-*', 'Control+Shift', 'B',   'Editor.SelPos:prevchar',    '左へ選択(C-S-B)'),
    ('*-Editor-Main-*', 'Control+Shift', 'F',   'Editor.SelPos:nextchar',    '右へ選択(C-S-F)'),
    ('*-Editor-Main-*', 'Control+Shift', 'A',   'Editor.SelPos:linestart',   '行頭まで選択'),
    ('*-Editor-Main-*', 'Control+Shift', 'E',   'Editor.SelPos:lineend',     '行末まで選択'),
    ('*-Editor-Main-*', 'Control+Shift+Alt','P','Editor.SelPos:firstline',   '先頭行まで選択'),
    ('*-Editor-Main-*', 'Control+Shift+Alt','N','Editor.SelPos:lastline',    '最終行まで選択'),
    # Editor: Search
    ('*-Editor-Main-*', 'Alt',           'F',   'Editor.SearchMode:next',    '検索/置換'),
    # Editor: Save
    ('*-Editor-Main-*', 'Control',       'S',   'Editor.Editing.Save',       '保存'),
    # Editor: Multicursor
    ('*-Editor-Main-*', 'Control+Alt',   'UP',  'Request.Invoke.Default',    'マルチカーソル↑'),
    ('*-Editor-Main-*', 'Control+Alt',   'DOWN','Request.Invoke.Default',    'マルチカーソル↓'),
    ('*-Editor-Main-*', 'Control',       'D',   'Request.Invoke.Default',    '次の一致を選択'),
    # Editor: Bullet/Comment/Folding level
    ('*-Editor-Main-*', 'Control',       ':',   'Editor.Edit.NextBullet',    '次の行頭文字'),
    ('*-Editor-Main-*', 'Control+Shift', '*',   'Editor.Edit.PrevBullet',    '前の行頭文字'),
    ('*-Editor-Main-*', 'Control',       '/',   'Editor.Edit.NextComment',   '次のコメント形式'),
    ('*-Editor-Main-*', 'Control+Shift', '?',   'Editor.Edit.PrevComment',   '前のコメント形式'),
    ('*-Editor-Main-*', 'Control+Shift+Alt','I','Editor.Edit.FoldingInit',   'Folding追加'),
    ('*-Editor-Main-*', 'Control+Shift', 'I',   'Editor.Edit.FoldingUp',     'Foldingレベル↑'),
    ('*-Editor-Main-*', 'Control',       'I',   'Editor.Edit.FoldingDown',   'Foldingレベル↓'),
    # Editor: AutoComplete
    ('*-Editor-Main-*', 'Control',       'SPACE','Editor.AutoComplete.Suggest','補完'),
    # Editor: Request
    ('*-Editor-Main-*', 'Alt',           'ENTER','Request.Invoke.Default',   'リクエスト実行'),
    ('*-Editor-Main-*', 'Alt+Shift',     'ENTER','Request.Show.ContextMenu', 'コンテキストメニュー'),
    # Editor: Folding open/close
    ('*-Editor-Main-*', 'Alt',           'RIGHT','Editor.Folding.Open',      '折りたたみ展開'),
    ('*-Editor-Main-*', 'Alt',           'LEFT', 'Editor.Folding.Close',     '折りたたみ閉じる'),
    ('*-Editor-Main-*', 'Alt+Shift',     'RIGHT','Editor.Folding.OpenAllSibling','同レベル全展開'),
    ('*-Editor-Main-*', 'Alt+Shift',     'LEFT', 'Editor.Folding.CloseAllSibling','同レベル全閉じる'),
    ('*-Editor-Main-*', 'Alt',           'B',   'Editor.Folding.Close',      '折りたたみ閉じる(B)'),
    ('*-Editor-Main-*', 'Control+Shift', '}',   'Editor.Folding.Open',       '折りたたみ展開(})'),
    ('*-Editor-Main-*', 'Control+Shift', '{',   'Editor.Folding.Close',      '折りたたみ閉じる({)'),
    # Editor: Folding navigation
    ('*-Editor-Main-*', 'Alt',           'UP',  'Editor.CurPos:prevvisfolding','前の見出し(↑)'),
    ('*-Editor-Main-*', 'Alt',           'DOWN','Editor.CurPos:nextvisfolding','次の見出し(↓)'),
    ('*-Editor-Main-*', 'Alt+Shift',     'UP',  'Editor.CurPos:prevsibfolding','前の兄弟見出し'),
    ('*-Editor-Main-*', 'Alt+Shift',     'DOWN','Editor.CurPos:nextsibfolding','次の兄弟見出し'),
    ('*-Editor-Main-*', 'Alt+Shift',     'P',   'Editor.CurPos:prevsibfolding','前の兄弟見出し(P)'),
    ('*-Editor-Main-*', 'Alt+Shift',     'N',   'Editor.CurPos:nextsibfolding','次の兄弟見出し(N)'),
    ('*-Editor-Main-*', 'Alt+Shift',     'F',   'Editor.CurPos:firstsibfolding','最初の兄弟見出し'),
    ('*-Editor-Main-*', 'Alt+Shift',     'B',   'Editor.CurPos:lastsibfolding','最後の兄弟見出し'),
    # Editor: ExFold mode
    ('*-Editor-Main-*', 'Control',       'K',   'ExMode:ExFold',             'ExFoldモード'),
    # Editor: Menu/Misc
    ('*-Editor-Main-*', '',              'F1',  'App.Delegate',              'ヘルプ/メニュー'),
    ('*-Editor-Main-*', 'Control',       'G',   'App.Delegate',              'Go to line'),
    ('*-Editor-Main-*', 'Alt',           'H',   'Tool:next',                 '次のツール'),

    # ---- Table: Move ----
    ('*-Table-*-*', '',              'UP',   'Table.CurPos:prev',         '前の行'),
    ('*-Table-*-*', '',              'DOWN', 'Table.CurPos:next',         '次の行'),
    ('*-Table-*-*', 'Shift',         'UP',   'Table.CurPos:-10',          '10行前'),
    ('*-Table-*-*', 'Shift',         'DOWN', 'Table.CurPos:+10',          '10行後'),
    ('*-Table-*-*', 'Shift+Control', 'UP',   'Table.CurPos:first',        '最初の行'),
    ('*-Table-*-*', 'Shift+Control', 'DOWN', 'Table.CurPos:last',         '最後の行'),
    ('*-Table-*-*', 'Control',       'P',    'Table.CurPos:prev',         '前の行(C-P)'),
    ('*-Table-*-*', 'Control',       'N',    'Table.CurPos:next',         '次の行(C-N)'),
    ('*-Table-*-*', 'Control+Shift', 'P',    'Table.CurPos:-10',          '10行前(C-S-P)'),
    ('*-Table-*-*', 'Control+Shift', 'N',    'Table.CurPos:+10',          '10行後(C-S-N)'),
    ('*-Table-*-*', 'Control',       'A',    'Table.CurPos:first',        '最初の行(C-A)'),
    ('*-Table-*-*', 'Control',       'E',    'Table.CurPos:last',         '最後の行(C-E)'),
    ('*-Table-*-*', 'Alt',           'P',    'Table.CurPos:prev',         '前の行(A-P)'),
    ('*-Table-*-*', 'Alt',           'N',    'Table.CurPos:next',         '次の行(A-N)'),
    ('*-Table-*-*', 'Alt+Shift',     'P',    'Table.CurPos:-10',          '10行前(A-S-P)'),
    ('*-Table-*-*', 'Alt+Shift',     'N',    'Table.CurPos:+10',          '10行後(A-S-N)'),
    # Table: Sort
    ('*-Table-*-*', '',      'F1',  'Table.SortCol1.Rev',        '列1ソート'),
    ('*-Table-*-*', '',      'F2',  'Table.SortCol2.Rev',        '列2ソート'),
    ('*-Table-*-*', '',      'F3',  'Table.SortCol3.Rev',        '列3ソート'),
    ('*-Table-*-*', '',      'F4',  'Table.SortCol4.Rev',        '列4ソート'),
    ('*-Table-*-*', '',      'F5',  'Table.SortCol5.Rev',        '列5ソート'),
    ('*-Table-*-*', 'Shift', 'F1',  'Table.SortProp1.Rev',       'プロパティ1ソート'),
    ('*-Table-*-*', 'Shift', 'F2',  'Table.SortProp2.Rev',       'プロパティ2ソート'),
    ('*-Table-*-*', 'Shift', 'F3',  'Table.SortProp3.Rev',       'プロパティ3ソート'),
    ('*-Table-*-*', 'Shift', 'F4',  'Table.SortProp4.Rev',       'プロパティ4ソート'),
    ('*-Table-*-*', 'Shift', 'F5',  'Table.SortProp5.Rev',       'プロパティ5ソート'),
    # Table: Request
    ('*-Table-*-*', '',              'SPACE',  'Request.Invoke.Default',    'リクエスト実行'),
    ('*-Table-*-*', 'Shift',         'SPACE',  'Request.Show.ContextMenu',  'コンテキストメニュー'),
    ('*-Table-*-*', 'Control',       'SPACE',  'Request.Invoke.Default',    'リクエスト実行(C)'),
    ('*-Table-*-*', 'Shift+Control', 'SPACE',  'Request.Show.ContextMenu',  'コンテキストメニュー(C)'),
    ('*-Table-*-*', 'Alt+Control',   'SPACE',  'Request.Invoke.Default',    'リクエスト実行(A-C)'),
    ('*-Table-*-*', '', 'Selection_LEFT2',     'Request.Invoke.Default',    'ダブルクリック'),
    ('*-Table-*-*', '', 'RIGHT1',              'Request.Show.ContextMenu',  '右クリック'),
    ('*-Table-*-*', '', 'TableTitle_LEFT1',    'Request.Invoke.Default',    'タイトルクリック'),
    ('*-Table-*-*', '', 'TableTitle_RIGHT1',   'Request.Show.ContextMenu',  'タイトル右クリック'),
    # Table: Resource/Tool
    ('*-Table-*-*', 'Alt', 'R', 'Table.Resource:Thinktank',   'リソース切替'),
    ('*-Table-*-*', 'Alt', 'H', 'Tool:Main',                  'Mainツールへ'),

    # ---- WebView: Move ----
    ('*-WebView-*-*', '',              'UP',   'WebView.CurPos:prev',       '前の項目'),
    ('*-WebView-*-*', '',              'DOWN', 'WebView.CurPos:next',       '次の項目'),
    ('*-WebView-*-*', 'Shift+Control', 'UP',   'WebView.CurPos:first',      '最初の項目'),
    ('*-WebView-*-*', 'Shift+Control', 'DOWN', 'WebView.CurPos:last',       '最後の項目'),
    ('*-WebView-*-*', 'Alt',           'P',    'WebView.CurPos:prev',       '前の項目(A-P)'),
    ('*-WebView-*-*', 'Alt',           'N',    'WebView.CurPos:next',       '次の項目(A-N)'),
    ('*-WebView-*-*', 'Alt+Shift',     'P',    'WebView.CurPos:first',      '最初の項目(A-S-P)'),
    ('*-WebView-*-*', 'Alt+Shift',     'N',    'WebView.CurPos:last',       '最後の項目(A-S-N)'),
    # WebView: Request
    ('*-WebView-*-*', 'Alt',       'ENTER',         'Request.Invoke.Default',   'リクエスト実行'),
    ('*-WebView-*-*', 'Alt+Shift', 'ENTER',         'Request.Show.ContextMenu', 'コンテキストメニュー'),
    ('*-WebView-*-*', '',          'LEFT1',          'Request.Invoke.Default',   'クリック'),
    ('*-WebView-*-*', '',          'Selection_LEFT2','Request.Invoke.Default',   'ダブルクリック'),
    ('*-WebView-*-*', '',          'RIGHT1',         'Request.Show.ContextMenu', '右クリック'),
    # WebView: Keyword
    ('*-WebView-Keyword-*', '', 'ENTER', 'WebView.Keyword.Query', 'キーワードクエリ実行'),
    # WebView: Tool
    ('*-WebView-*-*', 'Alt', 'H', 'Tool:Main', 'Mainツールへ'),

    # ---- ExApp mode ----
    ('*-*-*-ExApp', '',              ';',  'Font.Size:up',              'フォント拡大'),
    ('*-*-*-ExApp', '',              '-',  'Font.Size:down',            'フォント縮小'),
    ('*-*-*-ExApp', '',              'Z',  'Style:zen',                 'Zenレイアウト'),
    ('*-*-*-ExApp', '',              'S',  'Style:standard',            'Standardレイアウト'),
    ('*-*-*-ExApp', '',              'R',  'Style:reset',               'レイアウトリセット'),
    ('*-*-*-ExApp', '',              'V',  'Voice.Input:next',          '音声入力切替'),
    ('*-*-*-ExApp', 'Shift+Alt',     'R',  'Memo.Renew',                'メモ更新'),
    ('*-*-*-ExApp', 'Shift+Control+Alt','R','AllCollection.Save',       '全コレクション保存'),
    ('*-Editor-Main-ExApp', '', 'M', 'Editor.Minimap:next',       'ミニマップ切替'),
    ('*-Editor-Main-ExApp', '', 'F', 'Editor.Wordwrap:next',      '折り返し切替'),
    ('*-Editor-Main-ExApp', '', 'N', 'Editor.LineNumber:next',    '行番号切替'),
    ('*-Editor-Main-ExApp', '', 'R', 'Editor.SearchRegex:next',   '検索RegEx'),
    ('*-Editor-Main-ExApp', '', 'W', 'Editor.SearchWholeWord:next','検索単語単位'),
    ('*-Editor-Main-ExApp', '', 'C', 'Editor.SearchCase:next',    '検索大小文字'),
    ('*-Editor-Main-ExApp', '', 'P', 'Editor.ReplaceKeepCap:next','置換大文字保持'),
    ('*-Editor-Main-ExApp', '', 'L', 'Editor.ReplaceInSel:next',  '選択範囲内置換'),

    # ---- ExFold mode ----
    ('*-*-*-ExFold', '',      'RIGHT', 'Folding.Open',               '展開'),
    ('*-*-*-ExFold', '',      'LEFT',  'Folding.Close',              '閉じる'),
    ('*-*-*-ExFold', '',      'O',     'Folding.OpenAll',            '全展開'),
    ('*-*-*-ExFold', '',      'C',     'Folding.CloseAll',           '全閉じる'),
    ('*-*-*-ExFold', '',      '1',     'Folding.CloseAll',           '全閉じる(1)'),
    ('*-*-*-ExFold', '',      '2',     'Folding.OpenLevel2',         'Lv2まで展開'),
    ('*-*-*-ExFold', '',      '3',     'Folding.OpenLevel3',         'Lv3まで展開'),
    ('*-*-*-ExFold', '',      '4',     'Folding.OpenLevel4',         'Lv4まで展開'),
    ('*-*-*-ExFold', '',      '5',     'Folding.OpenLevel5',         'Lv5まで展開'),
    ('*-*-*-ExFold', '',      'P',     'Editor.CurPos:prevfolding',  '前の折りたたみ'),
    ('*-*-*-ExFold', '',      'N',     'Editor.CurPos:nextfolding',  '次の折りたたみ'),
    ('*-*-*-ExFold', 'Shift', 'RIGHT', 'Folding.OpenAllSibling',     '同レベル全展開'),
    ('*-*-*-ExFold', 'Shift', 'LEFT',  'Folding.CloseAllSibling',    '同レベル全閉じる'),

    # ---- ExDateTime mode ----
    ('*-*-*-ExDateTime', '',      'Y', 'DateTime.Next1y',            '1年後'),
    ('*-*-*-ExDateTime', 'Shift', 'Y', 'DateTime.Prev1y',            '1年前'),
    ('*-*-*-ExDateTime', '',      'M', 'DateTime.Next1m',            '1ヶ月後'),
    ('*-*-*-ExDateTime', 'Shift', 'M', 'DateTime.Prev1m',            '1ヶ月前'),
    ('*-*-*-ExDateTime', '',      'D', 'DateTime.Next1d',            '1日後'),
    ('*-*-*-ExDateTime', 'Shift', 'D', 'DateTime.Prev1d',            '1日前'),
    ('*-*-*-ExDateTime', '',      'K', 'DateTime.Next1w',            '1週間後'),
    ('*-*-*-ExDateTime', 'Shift', 'K', 'DateTime.Prev1w',            '1週間前'),
    ('*-*-*-ExDateTime', '',      'T', 'DateTime.Format:next',       '次のフォーマット'),
    ('*-*-*-ExDateTime', 'Shift', 'T', 'DateTime.Format:prev',       '前のフォーマット'),
    ('*-*-*-ExDateTime', '',      'W', 'DateTime.Weekday',           '曜日トグル'),
    ('*-*-*-ExDateTime', '',      'J', 'DateTime.Time',              '時刻トグル'),

    # ---- ExPanel mode ----
    ('*-*-*-ExPanel',       'Alt',           'Q',   '(ExPanel).Mode:Table',      'ExPanel Tableモード'),
    ('*-*-*-ExPanel',       'Alt',           'W',   '(ExPanel).Mode:WebView',    'ExPanel WebViewモード'),
    ('*-*-*-ExPanel',       'Alt',           'E',   '(ExPanel).Mode:Editor',     'ExPanel Editorモード'),
    ('*-*-*-ExPanel',       'Alt',           'M',   '(ExPanel).Mode:next',       'ExPanel 次モード'),
    ('*-*-*-ExPanel',       'Alt+Shift',     'M',   '(ExPanel).Mode:prev',       'ExPanel 前モード'),
    ('*-*-*-ExPanel',       '',              ';',   '(ExPanel).Font.Size:up',    'ExPanelフォント拡大'),
    ('*-*-*-ExPanel',       '',              '-',   '(ExPanel).Font.Size:down',  'ExPanelフォント縮小'),
    ('*-Editor-Main-ExPanel','Alt',          'ENTER','Request.Invoke.Default',   'リクエスト実行'),
    ('*-Editor-Main-ExPanel','Alt+Shift',    'ENTER','Request.Show.ContextMenu', 'コンテキストメニュー'),
    ('*-Editor-Main-ExPanel','Alt',          'P',   '(ExPanel).Editor.CurPos:prevrequest','前のリクエスト'),
    ('*-Editor-Main-ExPanel','Alt',          'N',   '(ExPanel).Editor.CurPos:nextrequest','次のリクエスト'),
    ('*-Table-*-ExPanel',   '',              'UP',  '(ExPanel).Table.CurPos:prev','前の行'),
    ('*-Table-*-ExPanel',   '',              'DOWN','(ExPanel).Table.CurPos:next','次の行'),
    ('*-Table-*-ExPanel',   'Shift',         'UP',  '(ExPanel).Table.CurPos:prev10','10行前'),
    ('*-Table-*-ExPanel',   'Shift',         'DOWN','(ExPanel).Table.CurPos:next10','10行後'),
    ('*-Table-*-ExPanel',   'Shift+Control', 'UP',  '(ExPanel).Table.CurPos:first','最初の行'),
    ('*-Table-*-ExPanel',   'Shift+Control', 'DOWN','(ExPanel).Table.CurPos:last', '最後の行'),
    ('*-Table-*-ExPanel',   'Control',       'P',   '(ExPanel).Table.CurPos:prev','前の行(C-P)'),
    ('*-Table-*-ExPanel',   'Control',       'N',   '(ExPanel).Table.CurPos:next','次の行(C-N)'),
    ('*-Table-*-ExPanel',   'Control',       'A',   '(ExPanel).Table.CurPos:first','最初の行(C-A)'),
    ('*-Table-*-ExPanel',   'Control',       'E',   '(ExPanel).Table.CurPos:last', '最後の行(C-E)'),
    ('*-Table-*-ExPanel',   'Control+Shift', 'P',   '(ExPanel).Table.CurPos:prev10','10行前(C-S-P)'),
    ('*-Table-*-ExPanel',   'Control+Shift', 'N',   '(ExPanel).Table.CurPos:next10','10行後(C-S-N)'),
    ('*-Table-*-ExPanel',   'Alt',           'P',   '(ExPanel).Table.CurPos:prev','前の行(A-P)'),
    ('*-Table-*-ExPanel',   'Alt',           'N',   '(ExPanel).Table.CurPos:next','次の行(A-N)'),
    ('*-Table-*-ExPanel',   'Alt+Shift',     'P',   '(ExPanel).Table.CurPos:prev10','10行前(A-S-P)'),
    ('*-Table-*-ExPanel',   'Alt+Shift',     'N',   '(ExPanel).Table.CurPos:next10','10行後(A-S-N)'),
    ('*-Table-*-ExPanel',   'Control',       'SPACE','Request.Invoke.Default',   'リクエスト実行'),
    ('*-Table-*-ExPanel',   'Shift+Control', 'SPACE','Request.Show.ContextMenu', 'コンテキストメニュー'),
]

# ============================================================
# データ分析ヘルパー
# ============================================================

def get_mode(ctx):
    parts = ctx.split('-')
    return parts[1] if len(parts) > 1 else '*'

def get_exmode(ctx):
    parts = ctx.split('-')
    return parts[3] if len(parts) > 3 else '*'

def get_tool(ctx):
    parts = ctx.split('-')
    return parts[2] if len(parts) > 2 else '*'

# Build lookup: (mods, key, mode) -> [(action, desc, exmode)]
lookup = {}  # key=(mods, key, mode), value=list of (action, desc, exmode, ctx)

for ctx, mods, key, action, desc in KEYBINDINGS:
    mode = get_mode(ctx)
    exmode = get_exmode(ctx)
    k = (mods, key, mode)
    if k not in lookup:
        lookup[k] = []
    lookup[k].append((action, desc, exmode, ctx))

def get_binding(mods, key, mode):
    """指定mode(*含む)のバインディングを返す。ExModeを除く通常モードのみ"""
    results = []
    for m in [mode, '*']:
        k = (mods, key, m)
        if k in lookup:
            for action, desc, exmode, ctx in lookup[k]:
                # 通常モード（ExMode指定なし）のみ
                if exmode == '*':
                    results.append((action, desc))
    return results

def get_binding_exmode(mods, key, exmode_name):
    """ExMode専用のバインディングを返す"""
    results = []
    k = (mods, key, '*')
    if k in lookup:
        for action, desc, exmode, ctx in lookup[k]:
            if exmode == exmode_name:
                results.append((action, desc, ctx))
    return results


# ============================================================
# Workbook作成
# ============================================================
wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: モード比較表
# ============================================================
ws1 = wb.active
ws1.title = '①モード比較表'

# ヘッダー説明
ws1['A1'] = 'ThinktankWebCC キーバインディング比較表'
ws1['A1'].font = fnt(bold=True, size=14)
ws1['A1'].fill = fill(C['header_bg'])
ws1['A1'].font = Font(bold=True, size=14, color='FFFFFF', name='Meiryo UI')
ws1.merge_cells('A1:J1')

ws1['A2'] = '凡例: ■緑=3モード共通 ■黄=2モード共通 ■薄青=Editor専用 ■薄橙=Table専用 ■薄緑=WebView専用 ■灰=グローバル ■薄赤=MASK(無効化)'
ws1['A2'].font = fnt(size=8)
ws1.merge_cells('A2:J2')
ws1.row_dimensions[2].height = 14

# カラム定義
COLS = ['修飾キー', 'キー', 'Editor', 'Table', 'WebView', '共通度', '説明(Editor)', '説明(Table)', '説明(WebView)', '備考']
for i, c in enumerate(COLS):
    h_cell(ws1.cell(3, i+1), c)

ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 32
ws1.column_dimensions['D'].width = 32
ws1.column_dimensions['E'].width = 32
ws1.column_dimensions['F'].width = 10
ws1.column_dimensions['G'].width = 24
ws1.column_dimensions['H'].width = 24
ws1.column_dimensions['I'].width = 24
ws1.column_dimensions['J'].width = 20

# キー一覧を定義 (優先度順)
MOD_ORDER = {
    '':                   0,
    'Shift':              1,
    'Control':            2,
    'Alt':                3,
    'Control+Shift':      4,
    'Shift+Control':      4,
    'Alt+Shift':          5,
    'Control+Alt':        6,
    'Control+Shift+Alt':  7,
    'Control+Alt+Shift':  7,
    'Shift+Control+Alt':  7,
    'Alt+Shift+Control':  7,
    'Alt+Control':        8,
}

KEY_ORDER_BASE = [
    # 矢印
    'UP', 'DOWN', 'LEFT', 'RIGHT',
    # ファンクション
    'F1', 'F2', 'F3', 'F4', 'F5', 'F6', 'F7', 'F8', 'F9', 'F10', 'F11', 'F12',
    # 特殊
    'SPACE', 'ENTER', 'TAB', 'ESC', 'BACKSPACE', 'DELETE',
    'HOME', 'END', 'PAGE_UP', 'PAGE_DOWN',
    # アルファベット
    'A','B','C','D','E','F','G','H','I','J','K','L','M',
    'N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
    # 数字
    '0','1','2','3','4','5','6','7','8','9',
    # 記号
    ';','-',':','*','/','?','{','}','[',']','\\','_','=','+','.',',','`',
    # マウス等
    'LEFT1','Selection_LEFT2','RIGHT1',
    'TableTitle_LEFT1','TableTitle_RIGHT1',
    'PanelTitle_RIGHT1','StatusBar_RIGHT1',
]

def key_sort(combo):
    mods, key = combo
    m_order = MOD_ORDER.get(mods, 99)
    try:
        k_order = KEY_ORDER_BASE.index(key)
    except ValueError:
        k_order = 999
    return (m_order, k_order, key)

# 全キーコンボを収集
all_combos = set()
for ctx, mods, key, action, desc in KEYBINDINGS:
    exmode = get_exmode(ctx)
    # ExMode専用でないもの
    if exmode == '*':
        all_combos.add((mods, key))

sorted_combos = sorted(all_combos, key=key_sort)

row = 4
prev_mod = None
for mods, key in sorted_combos:
    # 修飾キーが変わったらセクション区切り
    if mods != prev_mod:
        ws1.cell(row, 1).value = f'── {mods if mods else "(なし)"} ──'
        ws1.cell(row, 1).fill = fill(C['section_bg'])
        ws1.cell(row, 1).font = fnt(bold=True, color='FFFFFF', size=9)
        ws1.merge_cells(f'A{row}:J{row}')
        ws1.row_dimensions[row].height = 14
        row += 1
        prev_mod = mods

    e_binds = get_binding(mods, key, 'Editor')
    t_binds = get_binding(mods, key, 'Table')
    w_binds = get_binding(mods, key, 'WebView')
    g_binds = get_binding(mods, key, '*')

    has_e = bool(e_binds)
    has_t = bool(t_binds)
    has_w = bool(w_binds)
    has_g = bool(g_binds) and not (has_e or has_t or has_w)

    # MASK判定
    def is_mask(binds):
        return any('NoAction' in a for a, d in binds)

    e_mask = is_mask(e_binds)
    t_mask = is_mask(t_binds)
    w_mask = is_mask(w_binds)

    # 共通度判定
    active = sum([has_e and not e_mask, has_t and not t_mask, has_w and not w_mask])

    # アクション文字列
    def action_str(binds):
        if not binds:
            return ''
        return '\n'.join(f"{a}" for a, d in binds[:2])

    def desc_str(binds):
        if not binds:
            return ''
        return ' / '.join(d for a, d in binds[:2])

    e_act = action_str(e_binds or g_binds)
    t_act = action_str(t_binds or g_binds)
    w_act = action_str(w_binds or g_binds)

    e_desc = desc_str(e_binds)
    t_desc = desc_str(t_binds)
    w_desc = desc_str(w_binds)

    if has_g:
        e_act = t_act = w_act = action_str(g_binds)
        e_desc = t_desc = w_desc = desc_str(g_binds)

    # 共通度文字列
    if has_g:
        common_str = 'Global'
        row_bg = C['global']
    elif active == 3:
        common_str = '●●● 全共通'
        row_bg = C['common3']
    elif active == 2:
        common_str = '●● 2共通'
        row_bg = C['common2']
    elif active == 1:
        if has_e and not e_mask:
            common_str = 'Editor'
            row_bg = C['editor']
        elif has_t and not t_mask:
            common_str = 'Table'
            row_bg = C['table']
        else:
            common_str = 'WebView'
            row_bg = C['webview']
    else:
        common_str = 'MASK'
        row_bg = C['mask']

    # セルに書き込み
    def cell_bg(has, mask, global_bg):
        if has_g:
            return C['global']
        if mask:
            return C['mask']
        if has:
            return row_bg
        return C['empty']

    ws1.cell(row, 1).value = mods if mods else '(なし)'
    ws1.cell(row, 1).fill = fill(C['global'] if has_g else 'F0F0F0')
    ws1.cell(row, 1).font = fnt(size=9)
    ws1.cell(row, 1).alignment = aln()
    ws1.cell(row, 1).border = bdr()

    ws1.cell(row, 2).value = key
    ws1.cell(row, 2).fill = fill(C['global'] if has_g else 'F0F0F0')
    ws1.cell(row, 2).font = fnt(bold=True, size=9)
    ws1.cell(row, 2).alignment = aln(h='center')
    ws1.cell(row, 2).border = bdr()

    for col, (act, has, mask) in enumerate([
        (e_act, has_e or has_g, e_mask),
        (t_act, has_t or has_g, t_mask),
        (w_act, has_w or has_g, w_mask),
    ], 3):
        bg = cell_bg(has, mask, row_bg)
        ws1.cell(row, col).value = act
        ws1.cell(row, col).fill = fill(bg)
        ws1.cell(row, col).font = fnt(size=8)
        ws1.cell(row, col).alignment = aln()
        ws1.cell(row, col).border = bdr()

    ws1.cell(row, 6).value = common_str
    ws1.cell(row, 6).fill = fill(row_bg)
    ws1.cell(row, 6).font = fnt(bold=(active==3), size=9)
    ws1.cell(row, 6).alignment = aln(h='center')
    ws1.cell(row, 6).border = bdr()

    for col, desc in [(7, e_desc), (8, t_desc), (9, w_desc)]:
        ws1.cell(row, col).value = desc
        ws1.cell(row, col).fill = fill('FAFAFA')
        ws1.cell(row, col).font = fnt(size=8)
        ws1.cell(row, col).alignment = aln()
        ws1.cell(row, col).border = bdr()

    ws1.row_dimensions[row].height = 18
    row += 1

ws1.freeze_panes = 'A4'


# ============================================================
# Sheet 2: Editor詳細
# ============================================================
ws2 = wb.create_sheet('②Editor詳細')

ws2.merge_cells('A1:G1')
ws2['A1'] = 'Editor モード キーバインディング詳細'
ws2['A1'].fill = fill(C['editor'])
ws2['A1'].font = Font(bold=True, size=12, color='000000', name='Meiryo UI')

cols2 = ['カテゴリ', '修飾キー', 'キー', 'アクション', '説明', 'ExMode', 'コンテキスト']
for i, c in enumerate(cols2):
    h_cell(ws2.cell(2, i+1), c)

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 38
ws2.column_dimensions['E'].width = 24
ws2.column_dimensions['F'].width = 14
ws2.column_dimensions['G'].width = 28

# カテゴリ順で整理
EDITOR_CATEGORIES = [
    ('MASK/無効化', lambda ctx, mods, key, action, desc: get_mode(ctx) in ['Editor', '*'] and 'NoAction' in action and get_exmode(ctx) == '*'),
    ('カーソル移動', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'Editor' and 'CurPos' in action and 'Sel' not in action and get_exmode(ctx) == '*'),
    ('選択', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'Editor' and 'SelPos' in action and get_exmode(ctx) == '*'),
    ('折りたたみ操作', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'Editor' and ('Folding' in action) and get_exmode(ctx) == '*'),
    ('テキスト編集', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'Editor' and ('Edit.' in action or 'Delegate' in action) and get_exmode(ctx) == '*'),
    ('検索/置換', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'Editor' and 'Search' in action and get_exmode(ctx) == '*'),
    ('保存/補完', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'Editor' and ('Save' in action or 'AutoComplete' in action) and get_exmode(ctx) == '*'),
    ('リクエスト', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'Editor' and 'Request' in action and get_exmode(ctx) == '*'),
    ('ExApp (Editor)', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'Editor' and get_exmode(ctx) == 'ExApp'),
    ('ExPanel (Editor)', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'Editor' and get_exmode(ctx) == 'ExPanel'),
]

row2 = 3
for cat_name, cat_filter in EDITOR_CATEGORIES:
    cat_rows = [(ctx, mods, key, action, desc) for ctx, mods, key, action, desc in KEYBINDINGS if cat_filter(ctx, mods, key, action, desc)]
    if not cat_rows:
        continue

    # Section header
    ws2.merge_cells(f'A{row2}:G{row2}')
    ws2.cell(row2, 1).value = f'▶ {cat_name}'
    ws2.cell(row2, 1).fill = fill(C['section_bg'])
    ws2.cell(row2, 1).font = Font(bold=True, size=10, color='FFFFFF', name='Meiryo UI')
    ws2.row_dimensions[row2].height = 16
    row2 += 1

    for ctx, mods, key, action, desc in cat_rows:
        exmode = get_exmode(ctx)
        bg = C['mask'] if 'NoAction' in action else (C['exmode'] if exmode != '*' else C['editor'])
        d_cell(ws2.cell(row2, 1), cat_name, bg='F8F8F8')
        d_cell(ws2.cell(row2, 2), mods if mods else '(なし)')
        d_cell(ws2.cell(row2, 3), key, bold=True, h='center')
        d_cell(ws2.cell(row2, 4), action, bg=bg)
        d_cell(ws2.cell(row2, 5), desc)
        d_cell(ws2.cell(row2, 6), exmode if exmode != '*' else '')
        d_cell(ws2.cell(row2, 7), ctx, color='888888', size=8)
        ws2.row_dimensions[row2].height = 16
        row2 += 1

ws2.freeze_panes = 'A3'


# ============================================================
# Sheet 3: Table詳細
# ============================================================
ws3 = wb.create_sheet('③Table詳細')

ws3.merge_cells('A1:G1')
ws3['A1'] = 'Table モード キーバインディング詳細'
ws3['A1'].fill = fill(C['table'])
ws3['A1'].font = Font(bold=True, size=12, color='000000', name='Meiryo UI')

for i, c in enumerate(cols2):
    h_cell(ws3.cell(2, i+1), c, bg=C['header_bg'])

for col, w in zip(['A','B','C','D','E','F','G'], [18,22,20,38,24,14,28]):
    ws3.column_dimensions[col].width = w

TABLE_CATEGORIES = [
    ('カーソル移動', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'Table' and 'CurPos' in action and get_exmode(ctx) == '*'),
    ('ソート', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'Table' and 'Sort' in action),
    ('リクエスト', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'Table' and 'Request' in action and get_exmode(ctx) == '*'),
    ('リソース/ツール', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'Table' and ('Resource' in action or 'Tool' in action) and get_exmode(ctx) == '*'),
    ('ExPanel (Table)', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'Table' and get_exmode(ctx) == 'ExPanel'),
]

row3 = 3
for cat_name, cat_filter in TABLE_CATEGORIES:
    cat_rows = [(ctx, mods, key, action, desc) for ctx, mods, key, action, desc in KEYBINDINGS if cat_filter(ctx, mods, key, action, desc)]
    if not cat_rows:
        continue
    ws3.merge_cells(f'A{row3}:G{row3}')
    ws3.cell(row3, 1).value = f'▶ {cat_name}'
    ws3.cell(row3, 1).fill = fill(C['section_bg'])
    ws3.cell(row3, 1).font = Font(bold=True, size=10, color='FFFFFF', name='Meiryo UI')
    ws3.row_dimensions[row3].height = 16
    row3 += 1
    for ctx, mods, key, action, desc in cat_rows:
        exmode = get_exmode(ctx)
        bg = C['exmode'] if exmode != '*' else C['table']
        d_cell(ws3.cell(row3, 1), cat_name, bg='F8F8F8')
        d_cell(ws3.cell(row3, 2), mods if mods else '(なし)')
        d_cell(ws3.cell(row3, 3), key, bold=True, h='center')
        d_cell(ws3.cell(row3, 4), action, bg=bg)
        d_cell(ws3.cell(row3, 5), desc)
        d_cell(ws3.cell(row3, 6), exmode if exmode != '*' else '')
        d_cell(ws3.cell(row3, 7), ctx, color='888888', size=8)
        ws3.row_dimensions[row3].height = 16
        row3 += 1

ws3.freeze_panes = 'A3'


# ============================================================
# Sheet 4: WebView詳細
# ============================================================
ws4 = wb.create_sheet('④WebView詳細')

ws4.merge_cells('A1:G1')
ws4['A1'] = 'WebView モード キーバインディング詳細'
ws4['A1'].fill = fill(C['webview'])
ws4['A1'].font = Font(bold=True, size=12, color='000000', name='Meiryo UI')

for i, c in enumerate(cols2):
    h_cell(ws4.cell(2, i+1), c, bg=C['header_bg'])

for col, w in zip(['A','B','C','D','E','F','G'], [18,22,20,38,24,14,28]):
    ws4.column_dimensions[col].width = w

WEBVIEW_CATEGORIES = [
    ('カーソル移動', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'WebView' and 'CurPos' in action),
    ('リクエスト', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'WebView' and 'Request' in action),
    ('キーワードクエリ', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'WebView' and 'Keyword' in action),
    ('ツール', lambda ctx, mods, key, action, desc: get_mode(ctx) == 'WebView' and 'Tool' in action),
]

row4 = 3
for cat_name, cat_filter in WEBVIEW_CATEGORIES:
    cat_rows = [(ctx, mods, key, action, desc) for ctx, mods, key, action, desc in KEYBINDINGS if cat_filter(ctx, mods, key, action, desc)]
    if not cat_rows:
        continue
    ws4.merge_cells(f'A{row4}:G{row4}')
    ws4.cell(row4, 1).value = f'▶ {cat_name}'
    ws4.cell(row4, 1).fill = fill(C['section_bg'])
    ws4.cell(row4, 1).font = Font(bold=True, size=10, color='FFFFFF', name='Meiryo UI')
    ws4.row_dimensions[row4].height = 16
    row4 += 1
    for ctx, mods, key, action, desc in cat_rows:
        exmode = get_exmode(ctx)
        bg = C['webview']
        d_cell(ws4.cell(row4, 1), cat_name, bg='F8F8F8')
        d_cell(ws4.cell(row4, 2), mods if mods else '(なし)')
        d_cell(ws4.cell(row4, 3), key, bold=True, h='center')
        d_cell(ws4.cell(row4, 4), action, bg=bg)
        d_cell(ws4.cell(row4, 5), desc)
        d_cell(ws4.cell(row4, 6), exmode if exmode != '*' else '')
        d_cell(ws4.cell(row4, 7), ctx, color='888888', size=8)
        ws4.row_dimensions[row4].height = 16
        row4 += 1

ws4.freeze_panes = 'A3'


# ============================================================
# Sheet 5: Global/ExMode詳細
# ============================================================
ws5 = wb.create_sheet('⑤Global_ExMode詳細')

ws5.merge_cells('A1:G1')
ws5['A1'] = 'Global & ExMode キーバインディング詳細'
ws5['A1'].fill = fill(C['global'])
ws5['A1'].font = Font(bold=True, size=12, color='000000', name='Meiryo UI')

for i, c in enumerate(cols2):
    h_cell(ws5.cell(2, i+1), c, bg=C['header_bg'])

for col, w in zip(['A','B','C','D','E','F','G'], [18,22,16,40,24,14,28]):
    ws5.column_dimensions[col].width = w

GLOBAL_CATEGORIES = [
    ('Global: パネル切替', lambda ctx, mods, key, action, desc: get_mode(ctx) == '*' and get_exmode(ctx) == '*' and 'Panel' in action and 'Mode' not in action and 'ExMode' not in action),
    ('Global: モード切替', lambda ctx, mods, key, action, desc: get_mode(ctx) == '*' and get_exmode(ctx) == '*' and 'Mode' in action),
    ('Global: ExMode起動', lambda ctx, mods, key, action, desc: get_mode(ctx) == '*' and get_exmode(ctx) == '*' and 'ExMode' in action),
    ('Global: 検索/その他', lambda ctx, mods, key, action, desc: get_mode(ctx) == '*' and get_exmode(ctx) == '*' and 'Mode' not in action and 'Panel' not in action and 'ExMode' not in action),
    ('ExApp (Global)', lambda ctx, mods, key, action, desc: get_mode(ctx) == '*' and get_exmode(ctx) == 'ExApp'),
    ('ExFold', lambda ctx, mods, key, action, desc: get_exmode(ctx) == 'ExFold'),
    ('ExDateTime', lambda ctx, mods, key, action, desc: get_exmode(ctx) == 'ExDateTime'),
    ('ExPanel (Global)', lambda ctx, mods, key, action, desc: get_mode(ctx) == '*' and get_exmode(ctx) == 'ExPanel'),
]

row5 = 3
for cat_name, cat_filter in GLOBAL_CATEGORIES:
    cat_rows = [(ctx, mods, key, action, desc) for ctx, mods, key, action, desc in KEYBINDINGS if cat_filter(ctx, mods, key, action, desc)]
    if not cat_rows:
        continue
    ws5.merge_cells(f'A{row5}:G{row5}')
    ws5.cell(row5, 1).value = f'▶ {cat_name}'
    ws5.cell(row5, 1).fill = fill(C['section_bg'])
    ws5.cell(row5, 1).font = Font(bold=True, size=10, color='FFFFFF', name='Meiryo UI')
    ws5.row_dimensions[row5].height = 16
    row5 += 1
    for ctx, mods, key, action, desc in cat_rows:
        exmode = get_exmode(ctx)
        if 'ExApp' in exmode:
            bg = C['exmode']
        elif 'ExFold' in exmode:
            bg = 'E8F0FF'
        elif 'ExDateTime' in exmode:
            bg = 'FFE8F0'
        elif 'ExPanel' in exmode:
            bg = 'F0FFE8'
        else:
            bg = C['global']
        d_cell(ws5.cell(row5, 1), cat_name, bg='F8F8F8')
        d_cell(ws5.cell(row5, 2), mods if mods else '(なし)')
        d_cell(ws5.cell(row5, 3), key, bold=True, h='center')
        d_cell(ws5.cell(row5, 4), action, bg=bg)
        d_cell(ws5.cell(row5, 5), desc)
        d_cell(ws5.cell(row5, 6), exmode if exmode != '*' else 'Global')
        d_cell(ws5.cell(row5, 7), ctx, color='888888', size=8)
        ws5.row_dimensions[row5].height = 16
        row5 += 1

ws5.freeze_panes = 'A3'


# ============================================================
# Sheet 6: 空きキー分析
# ============================================================
ws6 = wb.create_sheet('⑥空きキー分析')

ws6.merge_cells('A1:H1')
ws6['A1'] = '空きキー分析 - モード別未割当キー一覧（新規アサイン候補）'
ws6['A1'].fill = fill('4F4F4F')
ws6['A1'].font = Font(bold=True, size=12, color='FFFFFF', name='Meiryo UI')

ws6.merge_cells('A2:H2')
ws6['A2'] = '※ GlobalキーはすべてのモードでBLOCK。MASKキーは明示的に無効化済み。空きキーはモード固有のアクション割当候補。'
ws6['A2'].font = fnt(size=8)

cols6 = ['修飾キー', 'キー', 'Global', 'Editor現状', 'Table現状', 'WebView現状', '空き(Editor)', '空き(Table)', '空き(WebView)']
for i, c in enumerate(cols6):
    h_cell(ws6.cell(3, i+1), c)

for col, w in zip(['A','B','C','D','E','F','G','H','I'], [22,14,10,26,26,26,12,12,12]):
    ws6.column_dimensions[col].width = w

# アルファベット + よく使うキーの組み合わせを調査
CHECK_MODS = ['', 'Shift', 'Control', 'Alt', 'Control+Shift', 'Alt+Shift', 'Control+Alt', 'Control+Shift+Alt']
CHECK_KEYS_ALPHA = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
CHECK_KEYS_FUNC  = [f'F{i}' for i in range(1,13)]
CHECK_KEYS_SPEC  = ['SPACE', 'ENTER', 'UP', 'DOWN', 'LEFT', 'RIGHT', 'TAB', 'ESC']

row6 = 4
prev_m = None

for mods in CHECK_MODS:
    keys_to_check = CHECK_KEYS_ALPHA + CHECK_KEYS_FUNC + CHECK_KEYS_SPEC

    section_written = False
    for key in keys_to_check:
        g = get_binding(mods, key, '*')
        e = get_binding(mods, key, 'Editor')
        t = get_binding(mods, key, 'Table')
        w = get_binding(mods, key, 'WebView')

        has_g  = bool(g)
        has_e  = bool(e)
        has_t  = bool(t)
        has_w  = bool(w)

        e_mask = any('NoAction' in a for a, d in e)
        t_mask = any('NoAction' in a for a, d in t)
        w_mask = any('NoAction' in a for a, d in w)

        # 空きかどうか
        e_free = not has_e and not has_g and not e_mask
        t_free = not has_t and not has_g and not t_mask
        w_free = not has_w and not has_g and not w_mask

        # 空きがあるキーのみ表示
        if not (e_free or t_free or w_free):
            continue

        if not section_written:
            ws6.merge_cells(f'A{row6}:I{row6}')
            ws6.cell(row6, 1).value = f'── 修飾キー: {mods if mods else "(なし)"} ──'
            ws6.cell(row6, 1).fill = fill(C['section_bg'])
            ws6.cell(row6, 1).font = Font(bold=True, size=9, color='FFFFFF', name='Meiryo UI')
            ws6.row_dimensions[row6].height = 14
            row6 += 1
            section_written = True

        def short_action(binds):
            if not binds:
                return ''
            a, d = binds[0]
            if 'NoAction' in a:
                return '[MASK]'
            return a.split('.')[-1][:20] if '.' in a else a[:20]

        d_cell(ws6.cell(row6, 1), mods if mods else '(なし)')
        d_cell(ws6.cell(row6, 2), key, bold=True, h='center')

        # Global
        g_txt = short_action(g) if g else ''
        d_cell(ws6.cell(row6, 3), g_txt, bg=C['global'] if g else 'FFFFFF')

        # Editor
        e_bg = C['mask'] if e_mask else (C['editor'] if has_e else C['empty_e'] if e_free else 'FFFFFF')
        d_cell(ws6.cell(row6, 4), short_action(e) if e else ('MASK' if e_mask else ''), bg=e_bg)

        # Table
        t_bg = C['mask'] if t_mask else (C['table'] if has_t else C['empty_t'] if t_free else 'FFFFFF')
        d_cell(ws6.cell(row6, 5), short_action(t) if t else ('MASK' if t_mask else ''), bg=t_bg)

        # WebView
        w_bg = C['mask'] if w_mask else (C['webview'] if has_w else C['empty_w'] if w_free else 'FFFFFF')
        d_cell(ws6.cell(row6, 6), short_action(w) if w else ('MASK' if w_mask else ''), bg=w_bg)

        # 空き表示
        d_cell(ws6.cell(row6, 7), '★ 空き' if e_free else '', bg='DAEEFF' if e_free else 'FFFFFF', bold=e_free, color='0000AA' if e_free else '000000', h='center')
        d_cell(ws6.cell(row6, 8), '★ 空き' if t_free else '', bg='FFF0E8' if t_free else 'FFFFFF', bold=t_free, color='AA4400' if t_free else '000000', h='center')
        d_cell(ws6.cell(row6, 9), '★ 空き' if w_free else '', bg='F0FFF0' if w_free else 'FFFFFF', bold=w_free, color='006600' if w_free else '000000', h='center')
        ws6.row_dimensions[row6].height = 16
        row6 += 1

ws6.freeze_panes = 'A4'

# ============================================================
# 保存
# ============================================================
output_path = r'C:\Users\gogow\Documents\ThinktankWebCC\ThinktankWebCC_KeyBindings.xlsx'
wb.save(output_path)
print(f'Excel saved: {output_path}')
