#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_default_events_from_excel.py
ThinktankWebCC_KeyBindings.xlsx の②③④⑤シートから DefaultEvents.ts を生成するスクリプト

使い方:
    python generate_default_events_from_excel.py
    python generate_default_events_from_excel.py --xlsx ThinktankWebCC_KeyBindings.xlsx --out src/Controllers/DefaultEvents.ts
    python generate_default_events_from_excel.py --dry-run   # 標準出力のみ

シート列レイアウト（②Editor詳細 / ③Table詳細 / ④WebView詳細 / ⑤Global_ExMode詳細 共通）:
    A: カテゴリ
    B: 修飾キー  （例: "Control", "Alt+Shift", "(なし)" など）
    C: キー      （例: "ENTER", "S", "UP" など）
    D: アクション（Excel内の短縮形 → ACTION_MAP で完全なActionIdに変換）
    E: 説明      （日本語コメント）
    F: ExMode    （"ExApp", "ExPanel", "Global", 空なら通常モード）
    G: コンテキスト（例: "*-Editor-Main-*"）
"""

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("エラー: openpyxl が見つかりません。'pip install openpyxl' でインストールしてください。")
    sys.exit(1)

# ============================================================
# 対象シート
# ============================================================
TARGET_SHEETS = ['②Editor詳細', '③Table詳細', '④WebView詳細', '⑤Global_ExMode詳細']

# ============================================================
# アクション名変換テーブル
# Excel D列の短縮形 → DefaultEvents.ts で使う完全なActionId
#
# ルール:
#   - "NoAction"         → "Application.Command.NoAction"
#   - "App.Delegate"     → "Application.Command.Delegate"
#   - "Editor.CurPos:*"  → "(Panel).Editor.CurPos:*"
#   - "Editor.SelPos:*"  → "(Panel).Editor.SelPos:*"
#   - "Editor.SearchMode:*" → "(Panel).Editor.SearchMode:*"
#   - "Editor.SearchRegex:*" → "(Panel).Editor.SearchRegex:*"
#   - "Editor.SearchWholeWord:*" → "(Panel).Editor.SearchWholeWord:*"
#   - "Editor.SearchCase:*" → "(Panel).Editor.SearchCaseSensitive:*"
#   - "Editor.SearchCcase:*" → "(Panel).Editor.SearchCaseSensitive:*"  (typo対応)
#   - "Editor.ReplaceKeepCap:*" → "(Panel).Editor.ReplaceKeepCapitalize:*"
#   - "Editor.ReplaceInSel:*" → "(Panel).Editor.ReplaceInSelection:*"
#   - "Editor.Minimap:*" → "(ExPanel).Editor.Minimap:*"  (ExApp)
#   - "Editor.Wordwrap:*" → "(ExPanel).Editor.Wordwrap:*"  (ExApp)
#   - "Editor.LineNumber:*" → "(ExPanel).Editor.LineNumber:*"  (ExApp)
#   - "Table.CurPos:*"   → "(Panel).Table.CurPos:*"
#   - "Table.Resource:*" → "(Panel).Table.Resource:*"
#   - "WebView.CurPos:*" → "(Panel).WebView.CurPos:*"
#   - "Panel:*"          → "Application.Current.Panel:*"
#   - "Mode:*"           → "Application.Current.Mode:*"
#   - "ExMode:*"         → "Application.Current.ExMode:*"
#   - "Tool:*"           → "Application.Current.Tool:*"
#   - "Font.Size:*"      → "Application.Font.Size:*"
#   - "Style:zen"        → "Application.Style.PanelRatio:zen"
#   - "Style:standard"   → "Application.Style.PanelRatio:standard"
#   - "Style:reset"      → "Application.Style.PanelRatio:reset"
#   - "Voice.Input:*"    → "Application.Voice.Input:*"
#   - "Memo.Renew"       → "Application.Memo.Renew"
#   - "AllCollection.Save" → "Application.AllCollection.Save"
#   - "DateTime.Next1y"  → "DateTime.Shift.Next1y"  (など)
#   - "(ExPanel).Mode:*" → "(ExPanel).Current.Mode:*"
#   - "Editor.CurPos:prevvisfolding" → "(Panel).Editor.CurPos:prevvisiblefolding"
#   - "Editor.CurPos:nextvisfolding" → "(Panel).Editor.CurPos:nextvisiblefolding"
# ============================================================

def transform_action(action_raw: str, exmode: str) -> str:
    """
    ExcelのD列アクション短縮形を、DefaultEvents.tsで使う完全なActionIdに変換する。
    """
    a = action_raw.strip()

    # --- 完全一致の変換（優先度高） ---
    EXACT = {
        'NoAction':           'Application.Command.NoAction',
        'App.Delegate':       'Application.Command.Delegate',
        'Memo.Renew':         'Application.Memo.Renew',
        'AllCollection.Save': 'Application.AllCollection.Save',
        'WebView.OpenSearch': 'WebView.OpenSearch',
        # DateTime (短縮形 → Shift付き)
        'DateTime.Next1y':    'DateTime.Shift.Next1y',
        'DateTime.Prev1y':    'DateTime.Shift.Prev1y',
        'DateTime.Next1m':    'DateTime.Shift.Next1m',
        'DateTime.Prev1m':    'DateTime.Shift.Prev1m',
        'DateTime.Next1d':    'DateTime.Shift.Next1d',
        'DateTime.Prev1d':    'DateTime.Shift.Prev1d',
        'DateTime.Next1w':    'DateTime.Shift.Next1w',
        'DateTime.Prev1w':    'DateTime.Shift.Prev1w',
        'DateTime.Weekday':   'DateTime.ChangeDetail.Weekday',
        'DateTime.Time':      'DateTime.ChangeDetail.Time',
        'DateTime.Format:next': 'DateTime.ChangeFormat.Next',
        'DateTime.Format:prev': 'DateTime.ChangeFormat.Prev',
        # ExPanel Mode
        '(ExPanel).Mode:Table':   '(ExPanel).Current.Mode:Table',
        '(ExPanel).Mode:WebView': '(ExPanel).Current.Mode:WebView',
        '(ExPanel).Mode:Editor':  '(ExPanel).Current.Mode:Editor',
        '(ExPanel).Mode:next':    '(ExPanel).Current.Mode:next',
        '(ExPanel).Mode:prev':    '(ExPanel).Current.Mode:prev',
        # ExFold
        'Folding.Open':            'Editor.Folding.Open',
        'Folding.Close':           'Editor.Folding.Close',
        'Folding.OpenAll':         'Editor.Folding.OpenAll',
        'Folding.CloseAll':        'Editor.Folding.CloseAll',
        'Folding.OpenLevel2':      'Editor.Folding.OpenLevel2',
        'Folding.OpenLevel3':      'Editor.Folding.OpenLevel3',
        'Folding.OpenLevel4':      'Editor.Folding.OpenLevel4',
        'Folding.OpenLevel5':      'Editor.Folding.OpenLevel5',
        'Folding.OpenAllSibling':  'Editor.Folding.OpenAllSibling',
        'Folding.CloseAllSibling': 'Editor.Folding.CloseAllSibling',
        # Table resource / tool
        'Table.Resource:Thinktank': '(Panel).Table.Resource:Thinktank',
        'Tool:Main':    'Application.Current.Tool:Main',
        'Tool:next':    'Application.Current.Tool:next',
        # Editor folding moves (ExFold context)
        'Editor.CurPos:prevfolding': '(Panel).Editor.CurPos:prevfolding',
        'Editor.CurPos:nextfolding': '(Panel).Editor.CurPos:nextfolding',
        # visfolding (短縮名のtypo対応)
        'Editor.CurPos:prevvisfolding':  '(Panel).Editor.CurPos:prevvisiblefolding',
        'Editor.CurPos:nextvisfolding':  '(Panel).Editor.CurPos:nextvisiblefolding',
        'Editor.CurPos:prevsibfolding':  '(Panel).Editor.CurPos:prevsibfolding',
        'Editor.CurPos:nextsibfolding':  '(Panel).Editor.CurPos:nextsibfolding',
        'Editor.CurPos:firstsibfolding': '(Panel).Editor.CurPos:firstsibfolding',
        'Editor.CurPos:lastsibfolding':  '(Panel).Editor.CurPos:lastsibfolding',
        # ExApp Editor options  → (ExPanel).Editor.*
        'Editor.Minimap:next':         '(ExPanel).Editor.Minimap:next',
        'Editor.Wordwrap:next':        '(ExPanel).Editor.Wordwrap:next',
        'Editor.LineNumber:next':      '(ExPanel).Editor.LineNumber:next',
        'Editor.SearchRegex:next':     '(Panel).Editor.SearchRegex:next',
        'Editor.SearchWholeWord:next': '(Panel).Editor.SearchWholeWord:next',
        'Editor.SearchCase:next':      '(Panel).Editor.SearchCaseSensitive:next',
        'Editor.ReplaceKeepCap:next':  '(Panel).Editor.ReplaceKeepCapitalize:next',
        'Editor.ReplaceInSel:next':    '(Panel).Editor.ReplaceInSelection:next',
        # SearchMode
        'Editor.SearchMode:next': '(Panel).Editor.SearchMode:next',
    }
    if a in EXACT:
        return EXACT[a]

    # --- プレフィックス変換 ---

    # Panel:* → Application.Current.Panel:*
    if a.startswith('Panel:'):
        return 'Application.Current.Panel:' + a[len('Panel:'):]

    # Mode:* → Application.Current.Mode:*
    if a.startswith('Mode:'):
        return 'Application.Current.Mode:' + a[len('Mode:'):]

    # ExMode:* → Application.Current.ExMode:*
    if a.startswith('ExMode:'):
        return 'Application.Current.ExMode:' + a[len('ExMode:'):]

    # Tool:* → Application.Current.Tool:*
    if a.startswith('Tool:'):
        return 'Application.Current.Tool:' + a[len('Tool:'):]

    # Font.Size:* → Application.Font.Size:*
    if a.startswith('Font.Size:'):
        return 'Application.Font.Size:' + a[len('Font.Size:'):]

    # Style:zen/standard/reset → Application.Style.PanelRatio:*
    if a.startswith('Style:'):
        return 'Application.Style.PanelRatio:' + a[len('Style:'):]

    # Voice.Input:* → Application.Voice.Input:*
    if a.startswith('Voice.Input:'):
        return 'Application.Voice.Input:' + a[len('Voice.Input:'):]

    # Editor.Editing.* / Editor.Edit.* / Editor.AutoComplete.* などはそのまま
    # Editor.CurPos:* → (Panel).Editor.CurPos:*
    if a.startswith('Editor.CurPos:'):
        return '(Panel).Editor.CurPos:' + a[len('Editor.CurPos:'):]

    # Editor.SelPos:* → (Panel).Editor.SelPos:*
    if a.startswith('Editor.SelPos:'):
        return '(Panel).Editor.SelPos:' + a[len('Editor.SelPos:'):]

    # Table.CurPos:* → (Panel).Table.CurPos:*
    if a.startswith('Table.CurPos:'):
        return '(Panel).Table.CurPos:' + a[len('Table.CurPos:'):]

    # WebView.CurPos:* → (Panel).WebView.CurPos:*
    if a.startswith('WebView.CurPos:'):
        return '(Panel).WebView.CurPos:' + a[len('WebView.CurPos:'):]

    # Keyword.CurPos:* → (Panel).Keyword.CurPos:*
    if a.startswith('Keyword.CurPos:'):
        return '(Panel).Keyword.CurPos:' + a[len('Keyword.CurPos:'):]

    # Keyword.SelPos:* → (Panel).Keyword.SelPos:*
    if a.startswith('Keyword.SelPos:'):
        return '(Panel).Keyword.SelPos:' + a[len('Keyword.SelPos:'):]

    # それ以外はそのまま（既に完全形の場合：Request.*, WebView.*, Editor.Edit.* など）
    return a


# ============================================================
# ユーティリティ
# ============================================================

def normalize_mods(raw) -> str:
    if raw is None:
        return ''
    s = str(raw).strip()
    if s in ('(なし)', '-', ''):
        return ''
    return s

def normalize_key(raw) -> str:
    if raw is None:
        return ''
    return str(raw).strip()

def normalize_exmode(raw) -> str:
    if raw is None:
        return ''
    s = str(raw).strip()
    # "Global" は通常モードの *-*-*-* コンテキストを示すラベルで ExMode ではない
    if s in ('', 'Global'):
        return ''
    return s

def ts_escape(s: str) -> str:
    return s.replace('\\', '\\\\')

def format_add_event(context: str, mods: str, key: str, action: str, comment: str) -> str:
    c = ts_escape(context)
    m = ts_escape(mods)
    k = ts_escape(key)
    a = ts_escape(action)
    line = f"    AddEvent('{c}', '{m}', '{k}', '{a}');"
    if comment:
        truncated = comment[:45]
        line = f"{line:<84}  // {truncated}"
    return line


# ============================================================
# Excelからイベントデータを読み込む
# ============================================================

def load_events_from_sheet(ws) -> list:
    """
    1シートからイベント行を読み込む。
    返り値: (type, data) のリスト
      type == '__section__' → data = セクション名文字列
      type == '__blank__'   → data = None
      type == '__event__'   → data = (context, mods, key, action, desc)
    """
    events = []
    current_section = ''

    for i, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True), 3):
        # 全列None → 空行
        if all(v is None for v in row):
            events.append(('__blank__', None))
            continue

        cat_raw  = row[0] if len(row) > 0 else None  # A: カテゴリ
        mods_raw = row[1] if len(row) > 1 else None  # B: 修飾キー
        key_raw  = row[2] if len(row) > 2 else None  # C: キー
        act_raw  = row[3] if len(row) > 3 else None  # D: アクション
        desc_raw = row[4] if len(row) > 4 else None  # E: 説明
        exm_raw  = row[5] if len(row) > 5 else None  # F: ExMode
        ctx_raw  = row[6] if len(row) > 6 else None  # G: コンテキスト

        cat_str = str(cat_raw).strip() if cat_raw else ''

        # セクション行 (A列が "▶ ...")
        if cat_str.startswith('▶'):
            current_section = cat_str.lstrip('▶').strip()
            events.append(('__section__', current_section))
            continue

        # アクション空ならスキップ
        action_raw = str(act_raw).strip() if act_raw else ''
        if not action_raw:
            continue

        mods   = normalize_mods(mods_raw)
        key    = normalize_key(key_raw)
        exmode = normalize_exmode(exm_raw)
        desc   = str(desc_raw).strip() if desc_raw else ''

        # コンテキストはG列を優先。なければ空文字（自動生成しない方針に変更）
        ctx = str(ctx_raw).strip() if ctx_raw else ''
        if not ctx:
            # G列なし → 警告だけ出してスキップ or コメントアウト
            print(f"  [警告] コンテキスト未設定の行をスキップ: row={i}, action={action_raw}", file=sys.stderr)
            continue

        if not key:
            continue

        # アクション名を変換
        action = transform_action(action_raw, exmode)

        events.append(('__event__', (ctx, mods, key, action, desc, current_section)))

    return events


# ============================================================
# TypeScript コード生成
# ============================================================

def generate_ts(all_events: list) -> str:
    lines = []

    lines.append("import type { TTModels } from '../models/TTModels';")
    lines.append("import { TTEvent } from '../models/TTEvent';")
    lines.append("")
    lines.append("// このファイルは generate_default_events_from_excel.py によって自動生成されました。")
    lines.append("// 手動編集後に再生成する際は generate_keybindings_excel.py → Excel 編集 → このスクリプトの流れで行ってください。")
    lines.append("")
    lines.append("export function InitializeDefaultEvents(models: TTModels) {")
    lines.append("    const events = models.Events;")
    lines.append("")
    lines.append("    function AddEvent(context: string, mods: string, key: string, actionId: string) {")
    lines.append("        // カンマ区切りキーの展開")
    lines.append("        if (key.includes(',')) {")
    lines.append("            key.split(',').forEach(k => AddEvent(context, mods, k.trim(), actionId));")
    lines.append("            return;")
    lines.append("        }")
    lines.append("        const ev = new TTEvent();")
    lines.append("        ev.Context = context;")
    lines.append("        ev.Mods    = mods;")
    lines.append("        ev.Key     = key;")
    lines.append("        ev.Name    = actionId;")
    lines.append("        ev.ID      = `${context}|${mods}|${key}`;")
    lines.append("        events.AddItem(ev);")
    lines.append("    }")
    lines.append("")

    seen = set()
    total_written = 0
    total_dup = 0

    for sheet_title, events in all_events:
        lines.append(f"    // {'=' * 62}")
        lines.append(f"    // {sheet_title}")
        lines.append(f"    // {'=' * 62}")
        lines.append("")

        last_section = None

        for etype, edata in events:
            if etype == '__blank__':
                lines.append("")
                continue

            if etype == '__section__':
                sec = edata
                if sec != last_section:
                    if last_section is not None:
                        lines.append("    // #endregion")
                        lines.append("")
                    lines.append(f"    // #region {sec}")
                    last_section = sec
                continue

            # __event__
            ctx, mods, key, action, desc, _sec = edata

            # セクションの自動 open
            if _sec != last_section:
                if last_section is not None:
                    lines.append("    // #endregion")
                    lines.append("")
                if _sec:
                    lines.append(f"    // #region {_sec}")
                last_section = _sec

            # 重複チェック
            dedup_key = (ctx, mods, key, action)
            if dedup_key in seen:
                total_dup += 1
                lines.append(f"    // [重複スキップ] AddEvent('{ctx}', '{mods}', '{key}', '{action}');")
                continue
            seen.add(dedup_key)

            lines.append(format_add_event(ctx, mods, key, action, desc))
            total_written += 1

        if last_section is not None:
            lines.append("    // #endregion")
        lines.append("")

    lines.append("}")
    lines.append("")

    print(f"  → 生成: {total_written} 件, 重複スキップ: {total_dup} 件", file=sys.stderr)
    return '\n'.join(lines)


# ============================================================
# メイン
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description='ThinktankWebCC_KeyBindings.xlsx の②③④⑤シートから DefaultEvents.ts を生成します。'
    )
    parser.add_argument('--xlsx', default='ThinktankWebCC_KeyBindings.xlsx')
    parser.add_argument('--out',  default='src/Controllers/DefaultEvents.ts')
    parser.add_argument('--dry-run', action='store_true', help='ファイルに書き込まず標準出力のみ')
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    out_path  = Path(args.out)

    if not xlsx_path.exists():
        print(f"エラー: {xlsx_path} が見つかりません。", file=sys.stderr)
        sys.exit(1)

    print(f"読み込み中: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # シート確認
    print(f"シート一覧: {wb.sheetnames}")

    all_events = []
    for sname in wb.sheetnames:
        if sname not in TARGET_SHEETS:
            continue
        ws = wb[sname]
        print(f"  処理中: '{sname}' (max_row={ws.max_row})")
        events = load_events_from_sheet(ws)
        event_count = sum(1 for t, _ in events if t == '__event__')
        print(f"    → {event_count} 件のイベント行を読み込みました。")
        all_events.append((sname, events))

    if not all_events:
        print("エラー: 対象シートが見つかりませんでした。", file=sys.stderr)
        sys.exit(1)

    print("TypeScriptコードを生成中...")
    ts_code = generate_ts(all_events)

    if args.dry_run:
        sys.stdout.buffer.write(ts_code.encode('utf-8'))
    else:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        if out_path.exists():
            backup = out_path.with_suffix('.ts.bak')
            import shutil
            shutil.copy2(str(out_path), str(backup))
            print(f"バックアップ: {backup}")
        out_path.write_text(ts_code, encoding='utf-8')
        print(f"生成完了: {out_path}  ({len(ts_code.splitlines())} 行)")

    print("完了。")


if __name__ == '__main__':
    main()
